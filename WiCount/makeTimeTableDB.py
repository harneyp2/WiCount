import db
import sqlite3 as lite
from sqlite3 import OperationalError
import glob, os
import openpyxl
import numpy as np
import wicount
import datetime
from dateutil.parser import parse

def GetRoomNo(room):
    room_no = room[:1] + "-"+ room[1:2] + room [3:]
    #print ("room ", room_no)
    return room_no

def GetDay(i):
    return {
        1: "Mon",
        3: "Tue",
        5: "Wed",
        7: "Thu",
        9: "Fri",
    }.get(i, 1)

def GetWeekNo(week):
    date = week.split("-")[0]
    date = parse(date)
    week1 = date.isocalendar()
    week_nos = [str(week1[1]) + "/" + str(week1[0]), str(week1[1] +1) + "/" + str(week1[0])]
    #print(week_nos)
    return week_nos
   
def AddDetailsToTimeTable(line, weekNo):
    ''' add the details to the database along with the week number '''
    sqlvalues = []
    module_codes = []
    start_time = wicount.GetTime(line[0])
    #build sql string
    for i in range(1,len(line),2):
        db_values = [room_id, GetDay(i), start_time, weekNo, line[i], line[i+1]]
        if line[i] and line[i+1] and line[i+1] != "N/A":
            #print ("here")
            module = [line[i], line[i+1]]
            module_codes.append(module)
        sqlvalues.append(db_values)
    try:
        #print(sqlvalues)
        c.executemany('INSERT OR IGNORE INTO timetable VALUES (?,?,?,?,?,?)', sqlvalues)
        #print("done: ", sqlvalues) 
    except OperationalError:
        print ("Command skipped: ", sqlvalues)
    con.commit()
    return module_codes

def UpdateModuleTable(module_list, week_no):
    try:
        c.execute ("create table if not exists modules(module varchar(12), week_no varchar(7), \
                no_students int, PRIMARY KEY (module, week_no));")
    except OperationalError:
        print("Modules table couldn't be created")
    con.commit()
    
    print (module_list)
    # get unique values from the module list
    #http://stackoverflow.com/questions/13464152/typeerror-unhashable-type-list-when-using-built-in-set-function
    module_list = sorted(set(map(tuple, module_list)))
    sqlvalues = []
    
    for x in range(0, len(module_list)):
        if x != len(module_list)-1 and module_list[x][0] == module_list[x+1][0]:
            continue
        else:
            values = [module_list[x][0], week_no, module_list[x][1]]
            sqlvalues.append(values)
            print(sqlvalues)
    try:            
        c.executemany('INSERT OR IGNORE INTO modules VALUES (?,?,?)', sqlvalues)
        #print("done: ", sqlvalues) 
    except OperationalError:
        print ("Command skipped: ", sqlvalues)
    con.commit()
       
con = db.get_connection()
#con = lite.connect('wicount.sqlite3')
c=con.cursor()
# if the table doesn't exist create it.
try:
    c.execute ("create table if not exists timetable(room_id INTEGER NOT NULL, day varchar(3), \
            time time, week_no varchar(7), module varchar(12), no_students int, \
            PRIMARY KEY (room_id, day, time, week_no));")
except OperationalError:
    print("Timetable table couldn't be created")
con.commit()
      
# Got help from http://stackoverflow.com/questions/3964681/find-all-files-in-directory-with-extension-txt-in-python
os.chdir("timetable")

#set up hard codeing this will need to be passed in.
campus = "Belfield"
building = "Computer Science"
for file in glob.glob("*.xlsx"):
    wb = openpyxl.load_workbook(file)
    sheetnames = wb.get_sheet_names()
    sqlvalues = []
    module_list = []
    for sheet in sheetnames:
        #print(sheet)
        if sheet != "All":
            
            sheetData = wb.get_sheet_by_name(sheet)
            
            #get the room capacity from the spread sheet
            capacity = sheetData.cell('C1','C1').value
            capacity = capacity.split()[2]
            #print(capacity)
            
            # get the week number for the timetable
            week1 = sheetData.cell('B1','B1').value
            week_nos = GetWeekNo(week1)
            week1 = week_nos[0]
            week2 = week_nos[1]
            
            room_details = [campus, building, GetRoomNo(sheet),int(capacity)] 
            room_id = wicount.GetRoomID(room_details)
            
            table = np.array([[cell.value for cell in col] for col in sheetData['A3':'K11']])
            #print(table)
            for line in table:
                module_list.extend(AddDetailsToTimeTable(line, week1))
            
            table = np.array([[cell.value for cell in col] for col in sheetData['M3':'W11']])
            #print(table)
            for line in table:
                module_list.extend(AddDetailsToTimeTable(line, week2))
    
    UpdateModuleTable(module_list, week1)
    
con.close() 